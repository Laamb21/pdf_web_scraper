#!/usr/bin/env python3

import os
import csv
from PyPDF2 import PdfReader
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill

def check_pdf_tagging(pdf_path: str) -> Dict:
    """Check if a PDF file has tagging."""
    try:
        with open(pdf_path, 'rb') as file:
            pdf = PdfReader(file)
            has_tags = '/StructTreeRoot' in pdf.trailer['/Root'] if '/Root' in pdf.trailer else False
            return {
                'filename': os.path.basename(pdf_path),
                'has_tags': has_tags,
                'page_count': len(pdf.pages)
            }
    except Exception as e:
        return {
            'filename': os.path.basename(pdf_path),
            'has_tags': False,
            'page_count': 0,
            'error': str(e)
        }

def process_directory(directory: str) -> List[Dict]:
    """Process all PDFs in the given directory."""
    try:
        # Ensure we have an absolute path
        directory = os.path.abspath(directory)
        
        # Check if drive is accessible
        drive = os.path.splitdrive(directory)[0]
        if not os.path.exists(drive):
            raise FileNotFoundError(f"Drive not found or not accessible: {drive}")
        
        # Comprehensive permission checks
        if not os.path.exists(directory):
            raise FileNotFoundError(f"Directory not found: {directory}")
            
        permissions_check = {
            'read': os.access(directory, os.R_OK),
            'write': os.access(directory, os.W_OK),
            'execute': os.access(directory, os.X_OK)
        }
        
        if not permissions_check['read']:
            print(f"Permission check results for {directory}:")
            print(f"Read: {'Yes' if permissions_check['read'] else 'No'}")
            print(f"Write: {'Yes' if permissions_check['write'] else 'No'}")
            print(f"Execute: {'Yes' if permissions_check['execute'] else 'No'}")
            raise PermissionError(f"Insufficient permissions for directory: {directory}")
            
        results = []
        for filename in os.listdir(directory):
            if filename.lower().endswith('.pdf'):
                pdf_path = os.path.join(directory, filename)
                result = check_pdf_tagging(pdf_path)
                results.append(result)
        return results
    except Exception as e:
        print(f"Error processing directory {directory}: {e}")
        return []

def generate_csv_report(results: List[Dict], output_file: str):
    """Generate a formatted CSV report of the results."""
    # Define columns and their display names
    fieldnames = {
        'filename': 'PDF Filename',
        'has_tags': 'Tagged Status',
        'page_count': 'Number of Pages',
        'error': 'Error Messages'
    }
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames.keys())
        
        # Write custom header with formatted column names
        writer.writerow(fieldnames)
        
        # Write data with formatted values
        for result in results:
            formatted_result = {
                'filename': result['filename'],
                'has_tags': 'Yes' if result['has_tags'] else 'No',
                'page_count': str(result['page_count']) if result['page_count'] > 0 else 'N/A',
                'error': result.get('error', 'None')
            }
            writer.writerow(formatted_result)

def generate_excel_report(results: List[Dict], output_file: str):
    """Generate a formatted Excel report of the results."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF Tag Report"

    # Define columns and their display names
    fieldnames = {
        'filename': 'PDF Filename',
        'has_tags': 'Tagged Status',
        'page_count': 'Number of Pages',
        'error': 'Error Messages'
    }

    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # Write headers
    for col, header in enumerate(fieldnames.values(), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data
    for row, result in enumerate(results, 2):
        formatted_result = {
            'filename': result['filename'],
            'has_tags': 'Yes' if result['has_tags'] else 'No',
            'page_count': result['page_count'] if result['page_count'] > 0 else 'N/A',
            'error': result.get('error', 'None')
        }
        for col, field in enumerate(fieldnames.keys(), 1):
            ws.cell(row=row, column=col, value=formatted_result[field])

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)

def generate_report(results: List[Dict], output_file: str):
    """Generate a report in either CSV or Excel format based on file extension."""
    file_ext = os.path.splitext(output_file)[1].lower()
    if file_ext == '.xlsx':
        generate_excel_report(results, output_file)
    else:  # Default to CSV
        generate_csv_report(results, output_file)

def verify_drive_access(path: str) -> bool:
    """Verify if the drive is accessible and has necessary permissions."""
    try:
        drive = os.path.splitdrive(path)[0]
        if not drive:
            return True  # Relative path, use current drive
            
        if not os.path.exists(drive):
            print(f"Drive not found: {drive}")
            return False
            
        test_file = os.path.join(drive, 'test_access.tmp')
        try:
            # Try to write a temporary file
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            return True
        except Exception as e:
            print(f"Drive access error: {str(e)}")
            return False
            
    except Exception as e:
        print(f"Error checking drive access: {str(e)}")
        return False

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Check PDFs for tagging')
    parser.add_argument('--dir', required=True, help='Directory containing PDFs to check')
    parser.add_argument('--output', default='pdf_tagging_report.csv', 
                       help='Output file (default: pdf_tagging_report.csv, use .xlsx extension for Excel format)')
    
    args = parser.parse_args()
    
    print(f"Processing PDFs in {args.dir}...")
    results = process_directory(args.dir)
    
    generate_report(results, args.output)
    print(f"\nReport generated: {args.output}")
    
    # Print summary
    total_pdfs = len(results)
    tagged_pdfs = sum(1 for r in results if r.get('has_tags', False))
    print(f"\nSummary:")
    print(f"Total PDFs processed: {total_pdfs}")
    print(f"PDFs with tags: {tagged_pdfs}")
    print(f"PDFs without tags: {total_pdfs - tagged_pdfs}")

if __name__ == "__main__":
    main()
